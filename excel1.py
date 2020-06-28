from openpyxl import load_workbook

wb = load_workbook('result\\demo.xlsx')
sheet = wb.active

sheet.cell(row=2, column=2).value = "Data Changed"
sheet['C2'] = "Data Changed Again"
wb.save('result\\demo.xlsx')
