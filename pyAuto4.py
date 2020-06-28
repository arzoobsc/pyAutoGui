import openpyxl
import pyautogui
import time
from openpyxl.utils import column_index_from_string

wb = openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames)
sheet = wb['Sheet1']
print(sheet.title)
head = sheet.cell(row=1, column=1)
print(head.value)
print(sheet['D3'].value)
active_sheet = wb.active

print(f"Number of Columns {sheet.max_column}")
print(f"Number of Rows {sheet.max_row}")

for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        c = sheet.cell(row=i, column=j).value
        print(c, end=" --- ")
    print()

# print(active_sheet.title)
# print(wb.get_sheet_names())

